#!/usr/bin/env python3
"""Aggregate all C64 optimization results into an xlsx workbook."""
import csv, os, sys
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# All measured data points (label, config, tps_list)
RESULTS = [
    ("Phase0 Control", "DSpark7, DSA-CP off, seq32", [789.80, 791.58, 806.76]),
    ("M0-1 1xC64", "Single client C64", [794.03, 814.13]),
    ("M0-1 2xC32", "Two clients C32 (aggregated)", [832.43, 807.33]),
    ("FusedMC2 DSpark7", "Fused MC2=1, DSpark7", [980.28, 978.06]),
    ("FusedMC2 DSpark3", "Fused MC2=1, DSpark3", [1100.08, 1186.37]),
    ("FusedMC2 DSpark3 v2", "Fused MC2=1, DSpark3 (3-run)", [1208.52, 1214.40, 1240.93]),
    ("FusedMC2 DSpark3 seq64", "Fused MC2=1, DSpark3, seq64, batch16k", [823.60, 852.88]),
    ("+MLAPO", "Fused MC2+DSpark3+MLAPO (2-run)", [1041.51, 1259.60]),
    ("+MLAPO v2", "Fused MC2+DSpark3+MLAPO (3-run)", [1231.95, 1220.39, 1254.97]),
    ("+DSA-CP (breakthrough)", "Fused MC2+DSpark3+MLAPO+DSA-CP", [1163.95, 1518.50, 1558.65]),
    ("5-round verify", "Fused MC2+DSpark3+MLAPO+DSA-CP (5-run)", [1523.54, 1502.34, 1583.17, 1505.61, 1540.39]),
    ("Final Acceptance", "Fused MC2+DSpark3+MLAPO+DSA-CP (3 formal)", [1495.57, 1524.63, 1548.44]),
]

# Profile data
PROFILE_OPS = [
    ("MoeDistributeDispatchV2", 11947400, 1333, 59.0),
    ("MoeDistributeCombineV2", 933169, 1333, 4.6),
    ("DynamicQuant", 653556, 8917, 3.2),
    ("MoeDistributeDispatchV2 (kernel)", 460715, 2322, 2.3),
    ("GroupedMatmulSwigluQuant", 409308, 2322, 2.0),
    ("Neg", 363432, 1806, 1.8),
    ("GroupedMatmulSwigluQuant (kernel)", 309652, 1634, 1.5),
    ("allgatherAicpuKernel", 252067, 1135, 1.2),
    ("MatMulV2", 250577, 6844, 1.2),
    ("QuantBatchMatmulV3", 248824, 10422, 1.2),
    ("QuantBatchMatmulV3 (variant)", 241721, 9116, 1.2),
    ("SparseAttnSharedkv", 177369, 2107, 0.9),
]

PROFILE_COMM = [
    ("hcom_allGather", 1700840, 5633, 42.7),
    ("hcom_reduceScatter", 1346608, 4128, 33.8),
    ("hcom_alltoallv", 863740, 903, 21.7),
    ("hcom_allReduce", 69462, 602, 1.7),
]

# Optimization path
OPT_PATH = [
    ("Phase 0", "Control group (DSpark7, DSA-CP off)", 796, "", "Baseline"),
    ("Phase 2", "+Fused MC2 (MoE comm fusion)", 979, "+23.0%", "Main effect: fuses dispatch+FFN+combine"),
    ("Phase 2", "DSpark7 -> DSpark3", 1221, "+53.4%", "Less spec overhead at high concurrency"),
    ("Phase 2", "+MLAPO", 1235, "+55.2%", "MLA Pool optimization"),
    ("Phase 2", "+DSA-CP on", 1523, "+91.3%", "Strong synergy with Fused MC2+DSpark3"),
]

wb = Workbook()

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
best_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

# Sheet 1: Optimization Path
ws1 = wb.active
ws1.title = "Optimization Path"
headers = ["Phase", "Configuration", "Mean TPS (tok/s)", "vs Baseline", "Key Finding"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, len(headers))
for i, (phase, config, tps, vs, finding) in enumerate(OPT_PATH, 2):
    ws1.cell(row=i, column=1, value=phase)
    ws17.cell(row=i, column=2, value=config) if False else ws1.cell(row=i, column=2, value=config)
    ws1.cell(row=i, column=3, value=tps)
    ws1.cell(row=i, column=4, value=vs)
    ws1.cell(row=i, column=5, value=finding)
    if "DSA-CP" in config and "on" in config:
        for c in range(1, 6):
            ws1.cell(row=i, column=c).fill = best_fill
auto_width(ws1)

# Sheet 2: All Results
ws2 = wb.create_sheet("All Results")
headers = ["Label", "Configuration", "Run", "TPS (tok/s)", "Mean", "Min", "Max", "CV%"]
for c, h in enumerate(headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, len(headers))
row = 2
for label, config, tps_list in RESULTS:
    import statistics
    mean = statistics.mean(tps_list)
    mn = min(tps_list)
    mx = max(tps_list)
    cv = statistics.stdev(tps_list) / mean * 100 if len(tps_list) > 1 else 0
    for j, tps in enumerate(tps_list, 1):
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=config)
        ws2.cell(row=row, column=3, value=j)
        ws2.cell(row=row, column=4, value=tps)
        if j == 1:
            ws2.cell(row=row, column=5, value=round(mean, 2))
            ws2.cell(row=row, column=6, value=round(mn, 2))
            ws2.cell(row=row, column=7, value=round(mx, 2))
            ws2.cell(row=row, column=8, value=round(cv, 2))
        if "Final" in label or "verify" in label:
            for c in range(1, 9):
                ws2.cell(row=row, column=c).fill = best_fill
        row += 1
auto_width(ws2)

# Sheet 3: Profile - Operator Statistics
ws3 = wb.create_sheet("Profile Op Statistics")
headers = ["Operator", "Total Time (us)", "Count", "Ratio (%)"]
for c, h in enumerate(headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, len(headers))
for i, (op, t, cnt, ratio) in enumerate(PROFILE_OPS, 2):
    ws3.cell(row=i, column=1, value=op)
    ws3.cell(row=i, column=2, value=t)
    ws3.cell(row=i, column=3, value=cnt)
    ws3.cell(row=i, column=4, value=ratio)
    if "MoeDistribute" in op:
        for c in range(1, 5):
            ws3.cell(row=i, column=c).fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
auto_width(ws3)

# Sheet 4: Profile - Communication
ws4 = wb.create_sheet("Profile Communication")
headers = ["Communication Op", "Total Time (us)", "Count", "Ratio (%)"]
for c, h in enumerate(headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, len(headers))
for i, (op, t, cnt, ratio) in enumerate(PROFILE_COMM, 2):
    ws4.cell(row=i, column=1, value=op)
    ws4.cell(row=i, column=2, value=t)
    ws4.cell(row=i, column=3, value=cnt)
    ws4.cell(row=i, column=4, value=ratio)
auto_width(ws4)

# Sheet 5: Final Config
ws5 = wb.create_sheet("Final Config")
configs = [
    ("tensor-parallel-size", "4", "TP4"),
    ("data-parallel-size", "4", "DP4"),
    ("enable-expert-parallel", "true", "EP16"),
    ("quantization", "ascend", "W8A8"),
    ("max-num-seqs", "32", "seq32"),
    ("max-num-batched-tokens", "8192", ""),
    ("block-size", "128", ""),
    ("gpu-memory-utilization", "0.90", ""),
    ("async-scheduling", "true", ""),
    ("speculative method", "dspark", ""),
    ("num_speculative_tokens", "3", "DSpark3 (high-concurrency optimal)"),
    ("enforce_eager", "true", "drafter eager mode"),
    ("cudagraph_mode", "FULL_DECODE_ONLY", ""),
    ("enable_fused_mc2", "1", "MoE comm fusion (MAIN EFFECT +23%)"),
    ("enable_mlapo", "1", "MLA Pool optimization"),
    ("enable_dsa_cp", "true", "DSA Compressor Pipeline (synergy +36%)"),
    ("multistream_overlap_shared_expert", "false", "Required by Fused MC2"),
    ("enable_npugraph_ex", "true", ""),
    ("enable_static_kernel", "false", ""),
    ("enable_cpu_binding", "true", ""),
    ("enable_prefix_caching", "false", ""),
]
headers = ["Parameter", "Value", "Notes"]
for c, h in enumerate(headers, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, len(headers))
for i, (param, val, note) in enumerate(configs, 2):
    ws5.cell(row=i, column=1, value=param)
    ws5.cell(row=i, column=2, value=val)
    ws5.cell(row=i, column=3, value=note)
    if "fused_mc2" in param or "dsa_cp" in param or "mlapo" in param:
        for c in range(1, 4):
            ws5.cell(row=i, column=c).fill = best_fill
auto_width(ws5)

outpath = sys.argv[1] if len(sys.argv) > 1 else "dsv4-c64-1280tps-result.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")
