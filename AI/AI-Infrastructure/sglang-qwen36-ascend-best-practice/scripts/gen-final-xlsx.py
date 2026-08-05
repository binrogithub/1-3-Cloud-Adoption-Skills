#!/usr/bin/env python3
"""Aggregate TP2/DP8 sticky-v2 + TP16 C1 results into a final xlsx.

Reads:
  V2  = .../sglang-bf16-nextn-tp2dp8-sticky-v2-YYYYMMDD/{scene}/concurrency_N/profile_export_aiperf.json
  TP16 = .../sglang-bf16-tp16-c1-YYYYMMDD/{scene}/profile_export_aiperf.json  (no concurrency subdir)

Writes 3 blocks: sglang-tp2dp8-sticky-v2 / sglang-tp16-c1 / sglang-BEST
  BEST = C1 from max(v2, tp16); C4/C8/C16 from v2.

Run on the test host (96). Adjust V2/TP16 paths if your date differs.
"""
import openpyxl, json, os, datetime

OUT_XLSX = "/home/qwen3.6-test/sglang-tp2dp8-final-result.xlsx"
today = datetime.date.today().strftime("%Y%m%d")

V2 = "/home/qwen3.6-test/sglang-bf16-nextn-tp2dp8-sticky-v2-" + today
TP16 = "/home/qwen3.6-test/sglang-bf16-tp16-c1-" + today

SCENES = [("chat", "Chat (128,256)"),
          ("coding", "Coding Agent (16384, 4096)"),
          ("sum", "Summarization (1024,128)")]

def get_per_user_v2(scene, c):
    """v2 sweep run: concurrency_N/profile_export_aiperf.json"""
    fp = os.path.join(V2, scene, "concurrency_%d" % c, "profile_export_aiperf.json")
    if not os.path.exists(fp):
        return None
    d = json.load(open(fp))
    v = d.get("output_token_throughput_per_user", {}).get("avg")
    return round(v, 2) if v is not None else None

def get_per_user_tp16(scene):
    """TP16 single-concurrency run: profile_export_aiperf.json (no concurrency_N subdir)"""
    fp = os.path.join(TP16, scene, "profile_export_aiperf.json")
    if not os.path.exists(fp):
        return None
    d = json.load(open(fp))
    v = d.get("output_token_throughput_per_user", {}).get("avg")
    return round(v, 2) if v is not None else None

# v2 (TP2/DP8 sticky) results: C1/C4/C8/C16
v2 = {}
for key, label in SCENES:
    v2[key] = {}
    for c in [1, 4, 8, 16]:
        v = get_per_user_v2(key, c)
        if v is not None:
            v2[key][c] = v
    print("v2 %-28s %s" % (label, v2[key]))

# TP16 C1 results
tp16_c1 = {}
for key, label in SCENES:
    v = get_per_user_tp16(key)
    if v is not None:
        tp16_c1[key] = v
    print("tp16 C1 %-28s %s" % (label, v))

# BEST: C1 = max(v2, tp16); C4/C8/C16 = v2
best = {}
for key, _ in SCENES:
    best[key] = {}
    for c in [1, 4, 8, 16]:
        vals = []
        if c in v2.get(key, {}):
            vals.append(v2[key][c])
        if c == 1 and key in tp16_c1:
            vals.append(tp16_c1[key])
        if vals:
            best[key][c] = max(vals)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

def write_block(start, name, sub, results, note=""):
    ws.cell(start, 1, "Output Token Throughput Per User")
    ws.cell(start, 2, "C1"); ws.cell(start, 3, "C4")
    ws.cell(start, 4, "C8"); ws.cell(start, 5, "C16")
    ws.cell(start, 6, name); ws.cell(start, 7, sub)
    for idx, (key, label) in enumerate(SCENES):
        r = start + 1 + idx
        ws.cell(r, 1, label)
        d = results.get(key, {})
        for ci, c in enumerate([1, 4, 8, 16]):
            ws.cell(r, 2 + ci, d.get(c, ""))
    ws.merge_cells(start_row=start, start_column=6, end_row=start + 3, end_column=6)
    ws.merge_cells(start_row=start, start_column=7, end_row=start + 3, end_column=7)
    if note: ws.cell(start, 9, note)

write_block(1, "sglang-tp2dp8-sticky-v2", "num-conv160/80,NEXTN", v2,
            "8x TP2/DP8 BF16+NEXTN sticky, num-conv 160/80, radix cache")
write_block(6, "sglang-tp16-c1", "TP16 single-engine", {k: {1: tp16_c1.get(k)} for k in tp16_c1},
            "TP16 16-die single engine, BF16+NEXTN, C1 only")
write_block(11, "sglang-BEST", "C1=TP16,C4+=TP2/DP8", best,
            "Hybrid: C1 from TP16 (16-die), C4/C8/C16 from TP2/DP8 sticky")
wb.save(OUT_XLSX)
print("saved: " + OUT_XLSX)
