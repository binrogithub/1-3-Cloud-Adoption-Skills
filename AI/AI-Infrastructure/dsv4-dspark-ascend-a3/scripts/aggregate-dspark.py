#!/usr/bin/env python3
"""Aggregate DSpark results + 3-way comparison (DSpark vs baseline vs Qwen) into xlsx."""
import json, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DSPARK = "/root/dsv4-test/raw-json-dspark"
BASELINE = "/root/dsv4-test/raw-json"
OUT = "/root/dsv4-test/dsv4-dspark-vs-baseline-result-20260807.xlsx"

SCEN = [("chat","Chat (128,256)"),("sum","Summarization (1024,128)"),("coding","Coding Agent (16384,4096)")]
CONC = [1,4,8,16]
QWEN = {"chat":[159.76,136.20,121.86,94.44],"sum":[148.32,135.09,120.29,92.14],"coding":[152.30,133.48,120.00,97.47]}

def load(basepath, d, c):
    f=os.path.join(basepath, d, f"c{c}.json")
    if not os.path.exists(f): return None
    j=json.load(open(f))
    return {
        "tput": j.get("output_token_throughput_per_user",{}).get("avg",0),
        "ttft_avg": j.get("time_to_first_token",{}).get("avg",0),
        "ttft_p50": j.get("time_to_first_token",{}).get("p50",0),
        "ttft_p99": j.get("time_to_first_token",{}).get("p99",0),
        "itl_avg": j.get("inter_token_latency",{}).get("avg",0),
        "itl_p50": j.get("inter_token_latency",{}).get("p50",0),
        "rlat": j.get("request_latency",{}).get("avg",0),
    }

wb = openpyxl.Workbook()
hf = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sf = Font(bold=True, size=11)
sfill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin = Side(border_style="thin", color="BFBFBF")
bd = Border(left=thin,right=thin,top=thin,bottom=thin)
ct = Alignment(horizontal="center", vertical="center")

def hdr(ws, row, n):
    for c in range(1,n+1):
        cell=ws.cell(row=row,column=c); cell.font=hf; cell.fill=hfill; cell.alignment=ct; cell.border=bd

# ===== Sheet 1: 3-way throughput =====
ws = wb.active; ws.title="3-way-Throughput"
ws.merge_cells("A1:G1")
ws["A1"]="DSpark vs Baseline (no spec) vs Qwen BF16+NEXTN — throughput tok/s/user"
ws["A1"].font=Font(bold=True,size=13); ws["A1"].alignment=ct
r=3
cols=["Scenario","C","Baseline","DSpark","DSpark/Baseline","Qwen BF16+NEXTN","DSpark/Qwen"]
for c,h in enumerate(cols,1): ws.cell(row=r,column=c,value=h)
hdr(ws,r,len(cols))
r+=1
for d,label in SCEN:
    for j,c in enumerate(CONC):
        b=load(BASELINE,d,c); s=load(DSPARK,d,c)
        bv=b["tput"] if b else 0; sv=s["tput"] if s else 0; qw=QWEN[d][j]
        sp=sv/bv if bv else 0; sq=sv/qw if qw else 0
        vals=[label if j==0 else "", f"C{c}", round(bv,2), round(sv,2), f"{sp:.2f}x", round(qw,2), f"{sq:.2f}x"]
        for c2,v in enumerate(vals,1):
            cell=ws.cell(row=r,column=c2,value=v); cell.border=bd; cell.alignment=ct
            if c2==5 and sp>1: cell.fill=green
            if c2==5 and sp<1: cell.fill=red
        r+=1
for c,w in enumerate([24,5,11,11,16,16,13],1): ws.column_dimensions[get_column_letter(c)].width=w

# ===== Sheet 2: DSpark ITL =====
ws2=wb.create_sheet("DSpark-ITL")
ws2.merge_cells("A1:E1"); ws2["A1"]="DSpark Inter-Token Latency (ms/token) vs Baseline"
ws2["A1"].font=Font(bold=True,size=13); ws2["A1"].alignment=ct
r=3
for c,h in enumerate(["Scenario","C","Baseline ITL","DSpark ITL","Reduction"],1): ws2.cell(row=r,column=c,value=h)
hdr(ws2,r,5); r+=1
for d,label in SCEN:
    for c in CONC:
        b=load(BASELINE,d,c); s=load(DSPARK,d,c)
        bi=b["itl_avg"] if b else 0; si=s["itl_avg"] if s else 0
        red=(1-si/bi) if bi else 0
        vals=[label if c==1 else "",f"C{c}",round(bi,2),round(si,2),f"-{red*100:.0f}%"]
        for c2,v in enumerate(vals,1):
            cell=ws2.cell(row=r,column=c2,value=v); cell.border=bd; cell.alignment=ct
        r+=1
for c,w in enumerate([24,5,13,13,12],1): ws2.column_dimensions[get_column_letter(c)].width=w

# ===== Sheet 3: DSpark TTFT =====
ws3=wb.create_sheet("DSpark-TTFT")
ws3.merge_cells("A1:E1"); ws3["A1"]="DSpark TTFT (ms)"
ws3["A1"].font=Font(bold=True,size=13); ws3["A1"].alignment=ct
r=3
for c,h in enumerate(["Scenario","C","TTFT avg","TTFT p50","TTFT p99"],1): ws3.cell(row=r,column=c,value=h)
hdr(ws3,r,5); r+=1
for d,label in SCEN:
    for c in CONC:
        s=load(DSPARK,d,c)
        if not s: continue
        vals=[label if c==1 else "",f"C{c}",round(s["ttft_avg"],1),round(s["ttft_p50"],1),round(s["ttft_p99"],1)]
        for c2,v in enumerate(vals,1):
            cell=ws3.cell(row=r,column=c2,value=v); cell.border=bd; cell.alignment=ct
        r+=1
for c,w in enumerate([24,5,10,10,10],1): ws3.column_dimensions[get_column_letter(c)].width=w

# ===== Sheet 4: Utilization comparison =====
ws4=wb.create_sheet("Utilization")
ws4.merge_cells("A1:D1"); ws4["A1"]="NPU Utilization — DSpark vs Baseline"
ws4["A1"].font=Font(bold=True,size=13); ws4["A1"].alignment=ct
r=3
for c,h in enumerate(["Metric","Baseline (no spec)","DSpark","Note"],1): ws4.cell(row=r,column=c,value=h)
hdr(ws4,r,4); r+=1
util=[
    ("Samples",2934,1835,""),
    ("AICore avg (%)",82.2,65.2,"DSpark lower - spec decode more efficient per step"),
    ("AICore max (%)",100,100,""),
    ("HBM avg (MB)",60342,51510,""),
    ("HBM avg (GB)",58.9,50.3,""),
    ("CPU avg (%)",6.7,8.4,""),
    ("Host Memory (GB)",864.0,863.4,""),
]
for k,bv,sv,note in util:
    ws4.cell(row=r,column=1,value=k).font=sf
    ws4.cell(row=r,column=1).fill=sfill; ws4.cell(row=r,column=1).border=bd
    ws4.cell(row=r,column=2,value=bv).border=bd; ws4.cell(row=r,column=2).alignment=ct
    ws4.cell(row=r,column=3,value=sv).border=bd; ws4.cell(row=r,column=3).alignment=ct
    ws4.cell(row=r,column=4,value=note).border=bd
    r+=1
ws4.column_dimensions["A"].width=20; ws4.column_dimensions["B"].width=18; ws4.column_dimensions["C"].width=14; ws4.column_dimensions["D"].width=45

# ===== Sheet 5: Config =====
ws5=wb.create_sheet("Config")
ws5.column_dimensions["A"].width=30; ws5.column_dimensions["B"].width=70
cfg=[
    ("Model","DeepSeek-V4-Flash-0731-w8a8 (284B, 13B activated, DSpark)"),
    ("Framework","vLLM 0.25.1 (quay.io/ascend/vllm-ascend:DeepSeekV4-flash-0731-a3)"),
    ("Speculative","DSpark, num_speculative_tokens=7, enforce_eager=true"),
    ("Quantization","W8A8 int8 (modelslim, --quantization ascend)"),
    ("Hardware","8 x Ascend 910 A3 = 16 die, 64GB HBM/die"),
    ("Topology","TP4 + DP4, expert-parallel, port 6697"),
    ("block-size","128"),
    ("max-model-len","131072"),
    ("Benchmark","aiperf 0.11.0, 3 scenarios x C1/C4/C8/C16"),
    ("Test Date","2026-08-07"),
]
for i,(k,v) in enumerate(cfg,1):
    ws5.cell(row=i,column=1,value=k).font=Font(bold=True)
    ws5.cell(row=i,column=2,value=v)

wb.save(OUT)
print(f"Saved: {OUT}")
print("Sheets:", wb.sheetnames)
